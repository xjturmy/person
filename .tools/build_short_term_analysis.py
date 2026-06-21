#!/usr/bin/env python3
"""
通用工具：从指定公司目录的 01–05 模块 CSV 提取「最近约一个月」数据，生成短期数据分析 xlsx。

使用方法：
    python build_short_term_analysis.py <公司目录路径> [参考日期] [窗口天数]

示例：
    # 使用默认参考日期（今天）和31天窗口
    python build_short_term_analysis.py ./01_新华保险
    
    # 指定参考日期和窗口
    python build_short_term_analysis.py ./01_新华保险 2026-04-12 31
    
    # 处理其他公司
    python build_short_term_analysis.py ./02_三美股份
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def parse_cell_date(s: str) -> date | None:
    s = (s or "").strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def file_base_key(stem: str) -> str:
    """同一指标多次导出时，按 stem 去重：去掉末尾 _YYYYMMDD_HHMMSS。"""
    parts = stem.rsplit("_", 2)
    if (
        len(parts) == 3
        and parts[1].isdigit()
        and len(parts[1]) == 8
        and parts[2].isdigit()
        and len(parts[2]) == 6
    ):
        return parts[0]
    return stem


def extract_company_name(base_path: Path) -> str:
    """从目录名提取公司名，如 '01_新华保险' -> '新华保险'"""
    name = base_path.name
    # 移除前导编号，如 "01_"、"02_"
    if re.match(r"^\d+_", name):
        name = re.sub(r"^\d+_", "", name)
    return name


def collect_csv_files(base_path: Path) -> list[Path]:
    """仅 01_估值分析 … 05_安全性分析，去重后每个指标保留最新导出文件。"""
    by_key: dict[str, Path] = {}
    for mod in ("01_估值分析", "02_盈利分析", "03_成长性分析", "04_现金流分析", "05_安全性分析"):
        d = base_path / mod
        if not d.is_dir():
            continue
        for p in sorted(d.glob("*.csv")):
            key = (mod, file_base_key(p.stem))
            prev = by_key.get(key)
            if prev is None or p.name > prev.name:
                by_key[key] = p
    return sorted(by_key.values(), key=lambda x: (x.parent.name, x.name))


def is_wide_quarterly(header: list[str]) -> bool:
    if not header or header[0].strip() != "日期":
        return False
    if len(header) < 2:
        return False
    return parse_cell_date(header[1]) is not None


def filter_wide_rows(rows: list[list[str]], header: list[str], ref_date: date, window_days: int) -> tuple[list[str], list[list[str]]]:
    """宽表：优先保留报告期落在最近 window_days 天内的列；若无则仅保留最新一期。"""
    period_cols: list[tuple[int, date]] = []
    for i, h in enumerate(header):
        if i == 0:
            continue
        d = parse_cell_date(h)
        if d is not None:
            period_cols.append((i, d))
    if not period_cols:
        return header, rows

    start = ref_date - timedelta(days=window_days)
    chosen_idx = [i for i, d in period_cols if d >= start]
    if not chosen_idx:
        latest = max(period_cols, key=lambda x: x[1])
        chosen_idx = [latest[0]]

    keep = [0] + sorted(chosen_idx)
    new_header = [header[i] for i in keep]
    new_rows = []
    for row in rows:
        new_rows.append([row[i] if i < len(row) else "" for i in keep])
    return new_header, new_rows


def filter_daily_rows(rows: list[list[str]], window_days: int) -> list[list[str]]:
    """日频：保留文件内最新日期往前 window_days 个自然日的行。"""
    dated: list[tuple[date, list[str]]] = []
    for row in rows:
        if not row:
            continue
        d = parse_cell_date(row[0])
        if d is not None:
            dated.append((d, row))
    if not dated:
        return rows
    max_d = max(t[0] for t in dated)
    cutoff = max_d - timedelta(days=window_days)
    return [row for d, row in dated if d >= cutoff]


def read_csv_raw(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        rows = list(r)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def strip_formula_equals(val: str) -> str:
    if val.startswith("="):
        return val[1:]
    return val


def main() -> None:
    parser = argparse.ArgumentParser(
        description="从公司目录的 01-05 模块 CSV 提取最近一个月数据生成 xlsx"
    )
    parser.add_argument(
        "company_dir",
        help="公司数据目录路径，如 './01_新华保险'"
    )
    parser.add_argument(
        "--ref-date",
        default=None,
        help="参考日期（用于计算窗口），格式 YYYY-MM-DD，默认为今天"
    )
    parser.add_argument(
        "--window-days",
        type=int,
        default=31,
        help="窗口天数（默认 31 天）"
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="输出文件路径（可选，默认在公司目录生成）"
    )

    args = parser.parse_args()

    base_path = Path(args.company_dir).resolve()
    if not base_path.is_dir():
        print(f"错误: 目录不存在: {base_path}", file=sys.stderr)
        sys.exit(1)

    # 解析参考日期
    if args.ref_date:
        try:
            ref_date = datetime.strptime(args.ref_date, "%Y-%m-%d").date()
        except ValueError:
            print(f"错误: 日期格式无效，应为 YYYY-MM-DD: {args.ref_date}", file=sys.stderr)
            sys.exit(1)
    else:
        ref_date = date.today()

    window_days = args.window_days
    company_name = extract_company_name(base_path)

    print(f"处理公司: {company_name}")
    print(f"参考日期: {ref_date}")
    print(f"窗口天数: {window_days}")
    print(f"数据源: {base_path}")

    files = collect_csv_files(base_path)

    if not files:
        print(f"警告: 在 {base_path} 中未找到 01-05 模块的 CSV 文件", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    summary = wb.create_sheet("说明与索引", 0)
    summary.append([f"{company_name} · 短期数据分析", "", "", ""])
    summary.append(["生成说明", "", "", ""])
    summary.append(
        [
            "日频 CSV（估值等）",
            f"保留文件内最新交易日至往前约 {window_days} 个自然日的全部行。",
            "",
            "",
        ]
    )
    summary.append(
        [
            "财报宽表（盈利/成长/现金流/安全等）",
            f"优先保留报告期截止日在「{ref_date}」往前约 {window_days} 天内的列；若该窗口内无财报列，则仅保留最新一期。",
            "",
            "",
        ]
    )
    summary.append(["源文件数（去重后）", len(files), "", ""])
    summary.append([])
    summary.append(["模块", "文件名", "类型", "说明"])

    title_font = Font(bold=True)

    for fp in files:
        mod = fp.parent.name
        header, body = read_csv_raw(fp)
        sheet_name = f"{mod[:8]}_{fp.stem[:15]}"
        sheet_name = re.sub(r'[\\/*?:\[\]]', "_", sheet_name)[:31]
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name[:28] + "_2"

        if is_wide_quarterly(header):
            h2, b2 = filter_wide_rows(body, header, ref_date, window_days)
            kind = "财报宽表·近月列"
            note = f"列数 {len(h2)}（含日期列）"
            ws = wb.create_sheet(sheet_name)
            ws.append(h2)
            for row in b2:
                ws.append([strip_formula_equals(c) for c in row])
        else:
            kind = "日频·近月行"
            filtered = filter_daily_rows(body, window_days)
            note = f"行数 {len(filtered)}（数据行，不含表头）"
            ws = wb.create_sheet(sheet_name)
            ws.append([strip_formula_equals(c) for c in header])
            for row in filtered:
                ws.append([strip_formula_equals(c) for c in row])

        ws.cell(row=1, column=1).font = title_font
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        summary.append([mod, fp.name, kind, note])

    # 确定输出路径
    if args.output:
        out_path = Path(args.output).resolve()
    else:
        out_path = base_path / f"{company_name}_短期数据分析.xlsx"

    wb.save(out_path)
    print(f"已写入: {out_path}")
    print(f"共处理 {len(files)} 个去重后的 CSV。")


if __name__ == "__main__":
    main()
